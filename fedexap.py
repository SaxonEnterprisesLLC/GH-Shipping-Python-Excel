# William Leonard, Saxon Enterprises LLC
# Contact: wfleonard@saxonenterprises.net 732-673-4260
# Fedex program for GH AR Shipping Data 
# Last Edit wfl 09-09-2022
# 09-09-2022    added csv_to_exel using pandas, cleaned up money issue 
#               added csv file extension check
#               added new GL entries to ledger dict
# 09-16-2022    added exit message

def conv_date(d):
    x = str(d)
    invoice = x[4:6] + x[6:8] + x[2:4]
    return invoice

def lookUpGL(gl):
    generalLedger = "9000-00-0000"
    ledger = {
            "8105004000" : "8105-00-4000",
            8105004000 : "8105-00-4000",
            "8105 00 4000" : "8105-00-4000",
            "8105-00-4000" : "8105-00-4000",
            "8105005000" : "8105-00-5000",
            8105005000 : "8105-00-5000",
            "8105-00-5000" : "8105-00-5000",
            "8105001000" : "8105-00-1000",
            8105001000 : "8105-00-1000",
            "8105-00-1000" : "8105-00-1000",
            "8105009000" : "8105-00-9000",
            8105009000 : "8105-00-9000",
            "8105003000" : "8105-00-3000",
            "8105-00-9000" : "8105-00-9000",
            "Ezekiel Edwards" : "8105-00-3000",
            "EZEKIEL EDWARDS" : "8105-00-3000",
            "Yesenia" : "8105-00-5000",
            "SAMPLE DEV - DUTY - SPRING" : "7810-00-1000",
            "SAMPLE DEV - DUTY - PREFALL" :	"7820-00-1000",
            "SAMPLE DEV - DUTY - RESORT" :  "7840-00-1000",
            "SAMPLE DEV - DUTY - FALL" :  "7830-00-1000",
            "PROD - SAMPLES DUTY - SPRING" : "7810-00-2000",
            "PROD - SAMPLES DUTY - PREFAL" : "7820-00-2000",
            "PROD - SAMPLES DUTY - FALL" : "7830-00-2000",
            "PROD - SAMPLES DUTY - RESORT" : "7840-00-2000",
            "PROD SAMPLES FREIGHT - SPRING" : "8010-00-2000",
            "PROD SAMPLES FREIGHT - PREFALL" : "8020-00-2000",
            "PROD SAMPLES FREIGHT - FALL" : "8030-00-2000",
            "PROD SAMPLES FREIGHT - RESORT" : "8040-00-2000",
            "WAREHOUSING - DUTY - HANDBAGS" : "8103-30-6000",
            "WAREHOUSING - DUTY - ECOM" : "8103-80-6000",
            "WAREHOUSING - DUTY - RETAIL" : "8103-90-6000",
            "WHSEING-  FREIGHT OUT - ECOM" : "8101-30-6000",
            "WHSEING-FREIGHT OUT-HANDBAG" : "8101-30-6000",
            "WHSING - FREIGHT OUT - WHOLESA" : "8101-00-6000",
            "WHSING- FREIGHT OUT - RETAIL" : "8101-90-6000",
            "WHSING-FREIGHT OUT - ECOM" : "8101-80-6000",
            "SHIPPING EXPENSE - DESIGN" : "8105-00-1000",
            "SHIPPING EXPENSE - E-COMMERCE" : "8105-00-4000",
            "SHIPPING EXPENSE - EXECUTIVE" : "8105-00-8000",
            "SHIPPING EXPENSE - G&A" : "8105-00-7000",
            "SHIPPING EXPENSE - MARKETING" : "8105-00-5000",
            "SHIPPING EXPENSE - PRODUCTION" : "8105-00-2000",
            "SHIPPING EXPENSE - RETAIL" : "8105-00-9000",
            "SHIPPING EXPENSE - SALES" : "8105-00-3000",
            "MESSENGER - DESIGN" : "8255-00-1000",
            "MESSENGER - ECOM" : "8255-00-4000",
            "MESSENGER - EXECUTIVE" : "8255-00-8000",
            "MESSENGER - G&A" : "8255-00-7000",
            "MESSENGER - MARKETING" : "8255-00-5000",
            "MESSENGER - PRODUCTION" : "8255-00-2000",
            "MESSENGER - RETAIL" : "8255-00-9000",
            "MESSENGER - SALES" : "8255-00-3000",
            "INVENTORY - DUTY - ACCESSORIES" : "1250-40-0000",
            "INVENTORY - DUTY - BAGS" : "1250-30-0000",
            "INVENTORY - DUTY - JEWELRY" : "1250-60-0000",
            "INVENTORY - DUTY - MENS" : "1250-50-0000",
            "INVENTORY - DUTY - RTW" : "1250-10-0000",
            "INVENTORY - DUTY - SHOES" : "1250-20-0000",
            "INVENTORY - FREIGHT - ACCESS" : "1240-40-0000",
            "INVENTORY - FREIGHT - BAGS" : "1240-30-0000",
            "INVENTORY - FREIGHT - JEWELRY" : "1240-60-0000",
            "INVENTORY - FREIGHT - MENS" : "1240-50-0000",
            "INVENTORY - FREIGHT - RTW" : "1240-10-0000",
            "INVENTORY - FREIGHT - SHOES" : "1240-20-0000"   
        }
    #print("gl ", gl)
    return ledger.get(gl, generalLedger)

def main():
    import time as t
    import csv_to_excel as csv
    import json
    from openpyxl import load_workbook

    terms="AD"
    #vendor_dhl=100311 
    vendor_fedex=100396
    #vendor_ups=100370
    #vender1="UPS"
    #vender2="DHL"
    vender3="FEDEX"
    numberFormat = '#,##0.00'
    interCompany = 1
    generalLedger = "9000-00-0000"

    #add menu logic to pick vender
    current_vender = vender3

    if current_vender == vender3:
        # Columns to use in conversion
        invDate=3
        invCol=4
        descCol=50
        glCol=52
        amountCol=12
        notesCol=108
        fedex_excel = "FedEx_Shipping_Feed.xlsx"
        fedex_sheet = "FEDEX-AP-Data"
        charge="Net Charge Amount"

        fedex_csv = input("Enter the FedEx shipping data file as .csv: ")
        
        while not fedex_csv.endswith(".csv"):
            print(f"\n\n{fedex_csv} is an invalid file name/extension")
            fedex_csv = input("Please input a .CSV file from FedEx: ")

        # Converts CSV file to an Excel .xlsx with proper datatypes (ie money)
        csv.csvToExcel(fedex_csv, fedex_excel, charge, current_vender)
        v_wb = load_workbook(filename=fedex_excel)
        
        v_ws = v_wb.active
        v_ws1 = v_wb.create_sheet(fedex_sheet)

        t_rows = v_ws.max_row
        t_cols = v_ws.max_column

        # create the master AR spreadsheet as a tab

        top = ["Vendor (VOUCHER_HEADER)",
        "Terms",
        "Invoice Date / Format : MMDDYY / Example : 010810",
        "Invoice Number (20 Characters Max)",
        "Description (50 Characters Max)",
        "Project Tracking",	
        "Pay To Bank",	
        "Factor",	
        "Inter Company",
        "General Ledger #Example:(VOUCHER_DETAILS)",
        "Amount Format / 2 Decimals / Example : 100.25",
        "Quantity",	
        "Division",
        "Season",
        "Year",
        "PO Number",
        "Notes (100 Characters Max)",
        "Delivery Period",	
        "Reference #",
        "Shipment #",
        "Subdivision",
        "Goods Or Services",
        "Destination Country",
        "Tax Rate",	
        "Tax Code",	
        "Style",	
        "Fabric",	
        "Length",
        "Color",
        "Cost Center",
        "Profit Center",
        "Project Act Tracking",
        "Summary Cost Code",
        "TempGL"]

        # Creates the header for the AP File

        y=1
        for header in top:
            v_ws1.cell(row=1, column=y).value = header
            y+=1

        #Create starting Invoice Number 
        invoiceNumber = "00000000"

        for r in range(2,t_rows+1):
            for c in range(1,t_cols):  
                if v_ws.cell(row=r, column=4).value != invoiceNumber:
                    v_ws1.cell(row=r, column=1).value = vendor_fedex                                        #Vendor Number
                    v_ws1.cell(row=r, column=2).value = terms                                               #Default Billing Terms
                    v_ws1.cell(row=r, column=3).value = conv_date(v_ws.cell(row=r, column=invDate).value)
                    v_ws1.cell(row=r, column=4).value = v_ws.cell(row=r, column=invCol).value               #Invoice Number
                    v_ws1.cell(row=r, column=5).value = v_ws.cell(row=r, column=descCol).value              #Description
                    v_ws1.cell(row=r, column=9).value = interCompany                                        #Inter Company
                    v_ws1.cell(row=r, column=10).value = generalLedger   #General Ledger
                    if v_ws.cell(row=r, column=amountCol).value is None:
                        v_ws1.cell(row=r, column=11).value = 0
                    else:
                        v_ws1.cell(row=r, column=11).value = v_ws.cell(row=r, column=amountCol).value           #Amount
                    v_ws1.cell(row=r, column=11).number_format = numberFormat   
                    v_ws1.cell(row=r, column=17).value = v_ws.cell(row=r, column=notesCol).value
                    v_ws1.cell(row=r, column=34).value = v_ws.cell(row=r, column=glCol).value          #Notes
                    invoiceNumber = v_ws.cell(row=r, column=4).value                                        #Incremented Invoice Number 
                else:
                    v_ws1.cell(row=r, column=5).value = v_ws.cell(row=r, column=descCol).value              #Description
                    v_ws1.cell(row=r, column=9).value = interCompany                            #Inter Company
                    v_ws1.cell(row=r, column=10).value = generalLedger                                      #General Ledger
                    if v_ws.cell(row=r, column=amountCol).value is None:
                        v_ws1.cell(row=r, column=11).value = 0
                    else:
                        v_ws1.cell(row=r, column=11).value = v_ws.cell(row=r, column=amountCol).value            #Amount
                    v_ws1.cell(row=r, column=11).number_format = numberFormat   
                    v_ws1.cell(row=r, column=17).value = v_ws.cell(row=r, column=notesCol).value
                    v_ws1.cell(row=r, column=34).value = v_ws.cell(row=r, column=glCol).value           #Notes
                    invoiceNumber = v_ws.cell(row=r, column=4).value                                        #Incremented Invoice Number 

        v_wb.save(fedex_excel)
    v_wb.close()

    workbook = load_workbook(filename=fedex_excel)
    workbook.active = 1
    sheet = workbook[fedex_sheet]

    for r in range(2,t_rows+1):
        for c in range(1,t_cols):
            sheet.cell(row=r, column=10).value = lookUpGL(sheet.cell(row=r, column=34).value)
            if sheet.cell(row=r, column=10).value == '9000-00-0000':
                sheet.cell(row=r, column=10).value = lookUpGL(sheet.cell(row=r, column=5).value)
    
    # removing the shipping codes column 
    sheet.delete_cols(34)
    workbook.save(fedex_excel)
    workbook.close()
    print()
    print(f"We have successfully created a FEDEX Excel file named {fedex_excel} with a Tab named {fedex_sheet}")
    t.sleep(3)

if __name__ == "__main__":
    main()
