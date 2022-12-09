# William Leonard, Saxon Enterprises LLC
# Contact: wfleonard@saxonenterprises.net 732-673-4260
# UPS program for GH AR Shipping Data 
# Last Edit wfl 10-21-2022
# 

def conv_date(d):
    dlist = list(d)
    dl = len(d)

    for x in range(dl):
        if dlist[1] == "-":
            dlist[1] = dlist[0]
            dlist[0] = "0"
            day = dlist[:2]
            mon = dlist[2:5]
            year = dlist[-2:]
            m=1
        elif dlist[2] == "-":
            day = dlist[:2]
            mon = dlist[3:6]
            year = dlist[-2:]
            m=2

    month = {
        "Jan":"01",
        "Feb":"02",
        "Mar":"03",
        "Apr":"04",
        "May":"05",
        "Jun":"06",
        "Jul":"07",
        "Aug":"08",
        "Sep":"09",
        "Oct":"10",
        "Nov":"11",
        "Dec":"12"
        }

    dd=""
    for y in dlist:
        dd += y

    if m==1:
        mon=month.get(dd[2:5], "XX")
    else:
        mon=month.get(dd[3:6], "XX")
    
    date = mon+dd[0:2]+dd[-2:]
    return date

def lookUpGL(gl):
    generalLedger = "9000-00-0000"
    ledger = {
            "SAMPLE DEV - DUTY - SPRING" : "7810-00-1000",
            "SAMPLE DEV - DUTY - PREFALL" :	"7820-00-1000",
            "SAMPLE DEV - DUTY - RESORT" :  "7840-00-1000",
            "SAMPLE DEV - DUTY - FALL" :  "7830-00-1000",
            "PROD - SAMPLES DUTY - SPRING" : "7810-00-2000",
            "PROD - SAMPLES DUTY - PREFAL" : "7820-00-2000",
            "PROD - SAMPLES DUTY - FALL" : "7830-00-2000",
            "PROD - SAMPLES DUTY - RESORT" : "7840-00-2000",
            "PROD SAMPLES FREIGHT - SPRING" : "8010-00-2000",
            "PROD SAMPLES FREIGHT - PREFALL" : "8020-00-2000",
            "PROD SAMPLES FREIGHT - FALL" : "8030-00-2000",
            "PROD SAMPLES FREIGHT - RESORT" : "8040-00-2000",
            "WAREHOUSING - DUTY - HANDBAGS" : "8103-30-6000",
            "WAREHOUSING - DUTY - ECOM" : "8103-80-6000",
            "WAREHOUSING - DUTY - RETAIL" : "8103-90-6000",
            "WHSEING-  FREIGHT OUT - ECOM" : "8101-30-6000",
            "WHSEING-FREIGHT OUT-HANDBAG" : "8101-30-6000",
            "WHSING - FREIGHT OUT - WHOLESA" : "8101-00-6000",
            "WHSING- FREIGHT OUT - RETAIL" : "8101-90-6000",
            "WHSING-FREIGHT OUT - ECOM" : "8101-80-6000",
            "SHIPPING EXPENSE - DESIGN" : "8105-00-1000",
            "SHIPPING EXPENSE - E-COMMERCE" : "8105-00-4000",
            "SHIPPING EXPENSE - EXECUTIVE" : "8105-00-8000",
            "SHIPPING EXPENSE - G&A" : "8105-00-7000",
            "SHIPPING EXPENSE - MARKETING" : "8105-00-5000",
            "SHIPPING EXPENSE - PRODUCTION" : "8105-00-2000",
            "SHIPPING EXPENSE - RETAIL" : "8105-00-9000",
            "SHIPPING EXPENSE - SALES" : "8105-00-3000",
            "MESSENGER - DESIGN" : "8255-00-1000",
            "MESSENGER - ECOM" : "8255-00-4000",
            "MESSENGER - EXECUTIVE" : "8255-00-8000",
            "MESSENGER - G&A" : "8255-00-7000",
            "MESSENGER - MARKETING" : "8255-00-5000",
            "MESSENGER - PRODUCTION" : "8255-00-2000",
            "MESSENGER - RETAIL" : "8255-00-9000",
            "MESSENGER - SALES" : "8255-00-3000",
            "INVENTORY - DUTY - ACCESSORIES" : "1250-40-0000",
            "INVENTORY - DUTY - BAGS" : "1250-30-0000",
            "INVENTORY - DUTY - JEWELRY" : "1250-60-0000",
            "INVENTORY - DUTY - MENS" : "1250-50-0000",
            "INVENTORY - DUTY - RTW" : "1250-10-0000",
            "INVENTORY - DUTY - SHOES" : "1250-20-0000",
            "INVENTORY - FREIGHT - ACCESS" : "1240-40-0000",
            "INVENTORY - FREIGHT - BAGS" : "1240-30-0000",
            "INVENTORY - FREIGHT - JEWELRY" : "1240-60-0000",
            "INVENTORY - FREIGHT - MENS" : "1240-50-0000",
            "INVENTORY - FREIGHT - RTW" : "1240-10-0000",
            "INVENTORY - FREIGHT - SHOES" : "1240-20-0000"   
        }
    #print("gl ", gl)
    return ledger.get(gl, generalLedger)

def main():
    import time as t
    import csv_to_excel as csv
    import json
    from openpyxl import load_workbook
    from openpyxl.styles import numbers

    terms="AD"
    vendor_ups=100370
    vender1="UPS"
    numberFormat = '#,##0.00'
    interCompany = 1
    generalLedger = "9000-00-0000"

    #add menu logic to pick vender
    current_vender = vender1

    if current_vender == vender1:
        # Columns to use in conversion
        invDate=3           #C
        invCol=2            #B
        descCol=11          #K
        glCol=7             #G Ref2
        amountCol=16        #P
        notesCol=6          #F Ref1
        tracking=4          #D
        #pubcharged=14      #N
        #incentives=15      #O
        ups_excel = "UPS_Shipping_Feed.xlsx"
        ups_sheet = "UPS-AP-Data"
        charge="Net Amount Due"

        ups_csv = input("Enter the UPS shipping data file as .csv: ")
        
        while not ups_csv.endswith(".csv"):
            print(f"\n\n{ups_csv} is an invalid file name/extension")
            ups_csv = input("Please input a .CSV file from UPS: ")

        # Converts CSV file to an Excel .xlsx with proper datatypes (ie money)
        csv.csvToExcel(ups_csv, ups_excel, charge, current_vender)
        v_wb = load_workbook(filename=ups_excel)
        
        v_ws = v_wb.active
        v_ws1 = v_wb.create_sheet(ups_sheet)

        t_rows = v_ws.max_row
        t_cols = v_ws.max_column

        # create the master AR spreadsheet as a tab

        top = ["Vendor (VOUCHER_HEADER)",
        "Terms",
        "Invoice Date / Format : MMDDYY / Example : 010810",
        "Invoice Number (20 Characters Max)",
        "Description (50 Characters Max)",
        "Project Tracking",	
        "Pay To Bank",	
        "Factor",	
        "Inter Company",
        "General Ledger #Example:(VOUCHER_DETAILS)",
        "Amount Format / 2 Decimals / Example : 100.25",
        "Quantity",	
        "Division",
        "Season",
        "Year",
        "PO Number",
        "Notes (100 Characters Max)",
        "Delivery Period",	
        "Reference #",
        "Shipment #",
        "Subdivision",
        "Goods Or Services",
        "Destination Country",
        "Tax Rate",	
        "Tax Code",	
        "Style",	
        "Fabric",	
        "Length",
        "Color",
        "Cost Center",
        "Profit Center",
        "Project Act Tracking",
        "Summary Cost Code",
        "TempGL"]

        # Creates the header for the AP File

        y=1
        for header in top:
            v_ws1.cell(row=1, column=y).value = header
            y+=1

        #Create starting Invoice Number 
        invoiceNumber = "000000000000"

        for r in range(2,t_rows+1):
            for c in range(1,t_cols):  
                if v_ws.cell(row=r, column=2).value != invoiceNumber:
                    v_ws1.cell(row=r, column=1).value = vendor_ups                                          #Vendor Number
                    v_ws1.cell(row=r, column=2).value = terms                                               #Default Billing Terms
                    v_ws1.cell(row=r, column=3).value = conv_date(v_ws.cell(row=r, column=invDate).value)
                    v_ws1.cell(row=r, column=4).value = v_ws.cell(row=r, column=invCol).value               #Invoice Number
                    v_ws1.cell(row=r, column=5).value = v_ws.cell(row=r, column=descCol).value              #Description
                    v_ws1.cell(row=r, column=9).value = interCompany                                        #Inter Company
                    v_ws1.cell(row=r, column=10).value = generalLedger
                    if v_ws.cell(row=r, column=amountCol).value == None:
                        v_ws1.cell(row=r, column=11).value = 0
                    else:                              
                        v_ws1.cell(row=r, column=11).value = v_ws.cell(row=r, column=amountCol).value
                    if v_ws1.cell(row=r, column=5).value == "Discounts":
                        v_ws1.cell(row=r, column=11).value *= -1
                    v_ws1.cell(row=r, column=11).number_format = numberFormat
                    v_ws1.cell(row=r, column=17).value = v_ws.cell(row=r, column=notesCol).value
                    v_ws1.cell(row=r, column=34).value = v_ws.cell(row=r, column=glCol).value                                                                                         #Notes
                    invoiceNumber = v_ws.cell(row=r, column=2).value                                        #Incremented Invoice Number 
                else:
                    v_ws1.cell(row=r, column=5).value = v_ws.cell(row=r, column=descCol).value              #Description
                    v_ws1.cell(row=r, column=9).value = interCompany                                        #Inter Company
                    v_ws1.cell(row=r, column=10).value = generalLedger                                      #General Ledger
                    if v_ws.cell(row=r, column=amountCol).value == None:
                        v_ws1.cell(row=r, column=11).value = 0
                    else:                              
                        v_ws1.cell(row=r, column=11).value = v_ws.cell(row=r, column=amountCol).value
                    if v_ws1.cell(row=r, column=5).value == "Discounts":
                        v_ws1.cell(row=r, column=11).value *= -1
                    v_ws1.cell(row=r, column=11).number_format = numberFormat
                    v_ws1.cell(row=r, column=17).value = v_ws.cell(row=r, column=notesCol).value
                    v_ws1.cell(row=r, column=34).value = v_ws.cell(row=r, column=glCol).value               #Notes
                    invoiceNumber = v_ws.cell(row=r, column=2).value                                        #Incremented Invoice Number 

        v_wb.save(ups_excel)
    v_wb.close()

    workbook = load_workbook(filename=ups_excel)
    workbook.active = 1
    sheet = workbook[ups_sheet]

    for r in range(2,t_rows+1):
        for c in range(1,t_cols):
            sheet.cell(row=r, column=10).value = lookUpGL(sheet.cell(row=r, column=34).value)
    
    # removing the shipping codes column 
    sheet.delete_cols(34)
    workbook.save(ups_excel)
    workbook.close()
    print()
    print(f"We have successfully created a UPS Excel file named {ups_excel} with a Tab named {ups_sheet}")
    t.sleep(2)

if __name__ == "__main__":
    main()
