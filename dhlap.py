# William Leonard, Saxon Enterprises LLC
# Contact: wfleonard@saxonenterprises.net 732-673-4260
# CSV to Excel program for GH AR Shipping Data 
# 10/14/22 Add new date conversion 

def conv_date(d):
    d = list(d)
    dl = len(d)

    for x in range(dl):
        if d[1] == "/":
            d[1] = d[0]
            d[0] = "0"

        if d[3] == "/":
            d[3] = d[2]
            d[2] = "0"

        if d[2] == "/" and d[4] == "/":
            d.remove(d[2])
            nl=len(d)
            d[3] = d[2]
            d[2] = "0"

        if d[2] == "/" and d[5] == "/":
            d.remove(d[2])
            nl=len(d) 

    nl=len(d)
    dd=""
    for y in d:
        dd += y
    
    date = dd[0:2]+dd[2:4]+dd[nl-2:]
    return date

def lookUpGL(gl):
    generalLedger = "9000-00-0000"
    ledger = {
            "8105004000" : "8105-00-4000",
            "8105 00 4000" : "8105-00-4000",
            "8105-00-4000" : "8105-00-4000",
            "8105005000" : "8105-00-5000",
            "8105-00-5000" : "8105-00-5000",
            "8105001000" : "8105-00-1000",
            "8105-00-1000" : "8105-00-1000",
            "8105009000" : "8105-00-9000",
            "8105-00-9000" : "8105-00-9000"
        }
    return ledger.get(gl, generalLedger)

def main():
    import time as t
    import csv_to_excel as csv
    import json
    from openpyxl import load_workbook


    terms="AD"
    vendor_dhl=100311 
    vendor_fedex=100396
    vendor_ups=100370
    vender1="UPS"
    vender2="DHL"
    vender3="FEDEX"
    numberFormat = '#,##0.00'
    interCompany = 1
    generalLedger = "9000-00-0000"

    #add menu logic to pick vender
    current_vender = vender2

    if current_vender == vender2:
        invDate=7
        invCol=2
        termsCol=8
        descCol=60
        amountCol=67
        notesCol=83
        dhl_excel = "DHL_Shipping_Feed.xlsx"
        dhl_sheet = "DHL-AP-Data"
        charge="Total Charge"

        dhl_csv = input("Enter the DHL shipping data file as .csv: ")
        while not dhl_csv.endswith(".csv"):
            print(f"\n\n{dhl_csv} is an invalid file name/extension")
            dhl_csv = input("Please input a .CSV file from DHL: ")

        # Converts CSV file to an Excel .xlsx with proper datatypes (ie money)
        csv.csvToExcel(dhl_csv, dhl_excel, charge, current_vender)

        v_wb = load_workbook(filename=dhl_excel)

        v_ws = v_wb.active
        v_ws1 = v_wb.create_sheet(dhl_sheet)

        t_rows = v_ws.max_row
        t_cols = v_ws.max_column
    
        # create the master AR spreadsheet as a tab

        top = ["Vendor (VOUCHER_HEADER)",
        "Terms",
        "Invoice Date / Format : MMDDYY / Example : 010810",
        "Invoice Number (20 Characters Max)",
        "Description (50 Characters Max)",
        "Project Tracking",	
        "Pay To Bank",	
        "Factor",	
        "Inter Company",
        "General Ledger #Example:(VOUCHER_DETAILS)",
        "Amount Format / 2 Decimals / Example : 100.25",
        "Quantity",	
        "Division",
        "Season",
        "Year",
        "PO Number",
        "Notes (100 Characters Max)",
        "Delivery Period",	
        "Reference #",
        "Shipment #",
        "Subdivision",
        "Goods Or Services",
        "Destination Country",
        "Tax Rate",	
        "Tax Code",	
        "Style",	
        "Fabric",	
        "Length",
        "Color",
        "Cost Center",
        "Profit Center",
        "Project Act Tracking",
        "Summary Cost Code"]

        # Creates the header for the AP File

        y=1
        for header in top:
            v_ws1.cell(row=1, column=y).value = header
            y+=1

        #Create starting Invoice Number 
        invoiceNumber = "00000000000"

        for r in range(2,t_rows+1):
            for c in range(1,t_cols):  
                if v_ws.cell(row=r, column=invCol).value != invoiceNumber:
                    v_ws1.cell(row=r, column=1).value = vendor_dhl
                    if v_ws.cell(row=r, column=termsCol).value != 15:
                        v_ws1.cell(row=r, column=2).value = terms
                    else: 
                        v_ws1.cell(row=r, column=2).value = v_ws.cell(row=r, column=termsCol).value                                        #Vendor Number                                              #Default Billing Terms
                    v_ws1.cell(row=r, column=3).value = conv_date(v_ws.cell(row=r, column=invDate).value)
                    v_ws1.cell(row=r, column=4).value = v_ws.cell(row=r, column=invCol).value               #Invoice Number
                    v_ws1.cell(row=r, column=5).value = v_ws.cell(row=r, column=descCol).value              #Description
                    v_ws1.cell(row=r, column=9).value = interCompany                                        #Inter Company
                    v_ws1.cell(row=r, column=10).value = generalLedger   #General Ledger 
                    if v_ws.cell(row=r, column=amountCol).value is None:
                        v_ws1.cell(row=r, column=11).value = 0
                    else:
                        v_ws1.cell(row=r, column=11).value = v_ws.cell(row=r, column=amountCol).value             #Amount
                    v_ws1.cell(row=r, column=11).number_format = numberFormat   
                    v_ws1.cell(row=r, column=17).value = v_ws.cell(row=r, column=notesCol).value            #Notes
                    invoiceNumber = v_ws.cell(row=r, column=invCol).value                                        #Incremented Invoice Number 
                else:
                    v_ws1.cell(row=r, column=5).value = v_ws.cell(row=r, column=descCol).value              #Description
                    v_ws1.cell(row=r, column=9).value = interCompany                            #Inter Company
                    v_ws1.cell(row=r, column=10).value = generalLedger                                      #General Ledger
                    if v_ws.cell(row=r, column=amountCol).value is None:
                        v_ws1.cell(row=r, column=11).value = 0
                    else:
                        v_ws1.cell(row=r, column=11).value = v_ws.cell(row=r, column=amountCol).value             #Amount
                    v_ws1.cell(row=r, column=11).number_format = numberFormat   
                    v_ws1.cell(row=r, column=17).value = v_ws.cell(row=r, column=notesCol).value            #Notes
                    invoiceNumber = v_ws.cell(row=r, column=invCol).value                                        #Incremented Invoice Number 

        v_wb.save(dhl_excel)
    v_wb.close()

    # Used in FEDEX change the GL Codes, need to close and reopen xlsx

    workbook = load_workbook(filename=dhl_excel)
    workbook.active = 1
    sheet = workbook[dhl_sheet]

    for r in range(2,t_rows+1):
        for c in range(1,t_cols):
            #print(lookUpGL(sheet.cell(row=r, column=c).value))
            sheet.cell(row=r, column=10).value = lookUpGL(sheet.cell(row=r, column=5).value)
    
    workbook.save(dhl_excel)
    workbook.close()
    print()
    print(f"We have successfully created a DHL Excel file named {dhl_excel} with a Tab named {dhl_sheet}")
    t.sleep(4)

if __name__ == "__main__":
    main()
